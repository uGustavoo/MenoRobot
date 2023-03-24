import PySimpleGUI as sg
from translate import Translator
from openpyxl import load_workbook


# Função da Tradução
def traduzir_arquivo(arquivo, lingua):
    try:
        wb = load_workbook(arquivo)
        ws = wb.active
        translator = Translator(to_lang=lingua)

        celula_total = ws.max_row * ws.max_column
        celula_atual = 0
        for linha in ws.iter_rows():
            for celula in linha:
                if celula.value:
                    celula.value = translator.translate(celula.value)
                    celula_atual += 1
                    porcentagem_progresso = int((celula_atual / celula_total) * 100)
                    janela["-PROGRESS-"].update(porcentagem_progresso)

        traduzir_arquivo = arquivo.split('.xlsx')[0] + f'_traduzido_{lingua}.xlsx'
        wb.save(traduzir_arquivo)
        return traduzir_arquivo
    except Exception as e:
        sg.popup_error(f"Ocorreu um erro: {str(e)}")
        return None


sg.theme('Default 1')

layout = [
    [sg.Text("Selecione o arquivo a ser traduzido:", font=('arial', 16), background_color='white')],
    [sg.Input(key="-FILE-", enable_events=True, visible=False),
     sg.FileBrowse("Escolher", target="-FILE-", size=(10, 1), button_color=('black', 'white'))],
    [sg.Text("Idioma de destino:", font=('arial', 16), background_color='white'),
     sg.Combo(["pt", "en", "ko"], default_value="pt", key="-LANG-", font=('arial', 14), size=(5, 1))],
    [sg.Button("Traduzir", disabled=True, font=('arial', 12), size=(10, 1), button_color=('white', '#c6194f'))],
    [sg.ProgressBar(100, orientation='h', size=(30, 10), key='-PROGRESS-')]
]

# Cria a janela
janela = sg.Window("Tradutor de Planilhas Excel", layout, size=(400, 200), element_justification='c',
                   icon=r'icon/icon.ico', background_color='white')

while True:
    evento, values = janela.read()
    if evento == sg.WIN_CLOSED:
        break
    if evento == "-FILE-":
        arquivo = values["-FILE-"]
        if arquivo.endswith(".xlsx"):
            janela.Element("Traduzir").Update(disabled=False)
    if evento == "Traduzir":
        lingua = values["-LANG-"]
        traduzir_arquivo = traduzir_arquivo(arquivo, lingua)
        if traduzir_arquivo:
            sg.popup(f"Planilha traduzida salva em {traduzir_arquivo}")
        janela.Element("Traduzir").Update(disabled=True)
        janela.Element("-PROGRESS-").Update(0)

janela.close()
